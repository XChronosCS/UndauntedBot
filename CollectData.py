import time

import gspread
import openpyxl
from googleapiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials

from gspread_credentials import *
from utilities import *

creds = ServiceAccountCredentials.from_json_keyfile_name('service_account_credentials.json',
                                                         scopes="https://www.googleapis.com/auth/documents.readonly")
service = build('docs', 'v1', credentials=creds)

gc = gspread.service_account_from_dict(credentials)

t1_start = time.perf_counter()

sh = gc.open("Data Get Test Sheet")
ec = gc.open("Porybot2 Encounters Sheet")
gp = gc.open("Guardian and Patronage Doc")
tt = gc.open("Test Town Tracker Sheet")

# Loading Worksheets for Primary Information Lookup
abilities = sh.worksheet("Abilities Data")
features = sh.worksheet("Features Data")
items = sh.worksheet("Inventory Data")
edges = sh.worksheet("Edges Data")
moves = sh.worksheet("Moves Data")
extras = sh.worksheet("Class Data")
misc = sh.worksheet("Misc Data")
habitat = gc.open("Data Habitat Areas").worksheet("Data")
encounters = ec.worksheet("Encounter Slots")
harvests = ec.worksheet("Harvest Slots")
events = ec.worksheet("Event Slots")
pokeedges = sh.worksheet("Tutoring/Breeding")
pokemon_data = sh.worksheet("Poke Data")
wander = gp.worksheet("Wander Events")
patrons = gp.worksheet("Patronage Tasks")
guardians = gp.worksheet("Guardian Table")
townevents = tt.worksheet("Town Data")
town_list = tt.worksheet("Town List")

worksheets = [("abilities", abilities, 1, 1, 3), ("features", features, 1, 1, 5), ("items", items, 2, 28, 29),
              ("edges", edges, 1, 1, 3), ("moves", moves, 1, 1, 9), ("mechanics", extras, 2, 1, 3),
              ("techniques", extras, 2, 4, 7),
              ("orders", features, 1, 6, 8), ("orders 2", features, 1, 9, 11), ("capabilities", misc, 1, 7, 8),
              ("keywords", misc, 1, 23, 24), ("statuses", misc, 1, 9, 11), ("maneuvers", moves, 1, 10, 15),
              ("books", misc, 1, 37, 43), ("weathers", misc, 1, 12, 13), ("affiliations", misc, 1, 14, 17),
              ("heritages", misc, 1, 18, 20), ("influences", misc, 1, 21, 22), ("pokeedges", pokeedges, 1, 1, 3),
              ("wanders", wander, 1, 1, 2), ("townevents", townevents, 1, 7, 8), ("uprisings", townevents, 1, 9, 10) ]
areas = [("encounters", encounters, 1), ("harvests", harvests, 1), ("events", events, 1)]
infodex = {}
worlddex = {}
eggdex = {}
bossdex = {
    "Guardians": {},
    "Patrons": {},
}

for worksheet in worksheets:

    # Prompt the user to enter the row number to use for the key names
    key_row_num = worksheet[2]
    start_col = worksheet[3]
    end_col = worksheet[4]

    # Get the data from the worksheet as a list of lists
    data = worksheet[1].get_all_values()

    # Get the key names from the specified row
    keys = data[key_row_num - 1][start_col - 1:end_col]

    # Initialize the dictionary to store the data
    data_dict = {}

    # Iterate over the rows of data and add them to the dictionary
    for i, row in enumerate(data):
        if i <= key_row_num - 1:
            continue  # Skip the row with the key names
        data_dict[row[start_col - 1]] = {keys[j]: row[j + start_col - 1] for j in range(len(keys))}

    infodex[worksheet[0]] = data_dict

infodex["orders"].update(infodex["orders 2"])


# Filling Out Worlddex
wb = openpyxl.load_workbook('Documents/Porybot2 Encounters Sheet.xlsx')

# Create a dictionary to store the data for all sheets

# Iterate over the sheets in the workbook
for sheet_name in wb.sheetnames:
    # Get the current sheet
    ws = wb[sheet_name]

    # Create a dictionary to store the data for the current sheet
    sheet_data = {}

    # Iterate over the columns in the worksheet
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=70):
        # Get the first cell in the column
        key = col[0].value

        # Create a nested dictionary for the column
        nested_dict = {}

        # Iterate over the cells in the column
        for cell in col[1:]:
            # Get the row number, cell value, and comment value
            row_num = cell.row - 1
            cell_val = cell.value
            comment_val = cell.comment.text if cell.comment else None
            hex_code = cell.fill.fgColor.rgb[2:].lower() if cell.fill and cell.fill.fgColor.type == "rgb" else None
            treasure_tag = "Major Treasure" if hex_code == "ffd966" else "Minor Treasure" if hex_code == "ffff00" else \
                "Guardian" if hex_code == "ff0000" else "Alpha Aberration" if hex_code == "#ea9999" else "Ignore"
            # Add the row number, cell value, and comment value to the nested dictionary
            if cell_val is not None:
                nested_dict[str(row_num)] = (cell_val, comment_val, treasure_tag)

        # Add the nested dictionary to the sheet data dictionary
        sheet_data[key] = nested_dict

    # Add the sheet data dictionary to the main dictionary
    worlddex[sheet_name] = sheet_data

# Filling Out Eggdex
eggbook = openpyxl.load_workbook('Documents/Data Undaunted Egg Rolls.xlsx')
# Get the current sheet
eggsheet = eggbook["Eggs"]

# Create a dictionary to store the data for the current sheet


# Iterate over the columns in the worksheet
for col in eggsheet.iter_cols(min_col=1, max_col=eggsheet.max_column, min_row=1, max_row=eggsheet.max_row):
    # Get the first cell in the column
    key = col[0].value

    # Create a nested dictionary for the column
    nested_dict = {}

    # Iterate over the cells in the column
    for cell in col[1:]:
        # Get the row number, cell value, and comment value
        row_num = cell.row
        cell_val = cell.value
        comment_val = create_item_list(cell.comment.text) if cell.comment else None
        if comment_val is not None:
            del comment_val[:2]

        # Add the row number, cell value, and comment value to the nested dictionary
        if cell_val is not None:
            nested_dict[str(row_num)] = (cell_val, comment_val)

    # Add the nested dictionary to the sheet data dictionary
    eggdex[key] = nested_dict

# Filling info in Bossdex

guardian_data = guardians.get_all_values()
keys = guardian_data[0][1:3]
for i, row in enumerate(guardian_data):
    if i <= 0:
        continue  # Skip the row with the key names
    bossdex["Guardians"][row[0]] = {keys[j]: row[j + 1] for j in range(len(keys))}


patronage_data = patrons.get_all_values()
for i, row in enumerate(patronage_data):
    if i <= 0:
        continue  # Skip the row with the key names
    if row[0] not in bossdex["Patrons"].keys():
        bossdex["Patrons"][row[0]] = {
            "Personality": None,
            "Minor": {},
            "Major": {},
            "Pact": {},
            "Task": {},
            "Custom": {},
        }
    bossdex["Patrons"][row[0]]["Personality"] = row[1]
    bossdex["Patrons"][row[0]][row[2]][row[3]] = [row[j] for j in range(4, 23) if row[j] != '']

t1_stop = time.perf_counter()
print("Elapsed time during the whole program in seconds:",
      t1_stop - t1_start)
